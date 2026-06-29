import os
import random
import sys
import time
from datetime import datetime

import openpyxl
from playwright.sync_api import sync_playwright, Page, TimeoutError as PWTimeout

# ── Sender info ───────────────────────────────────────────────────────────────
FIRST_NAME = os.environ.get("SENDER_FIRST_NAME", "")
LAST_NAME  = os.environ.get("SENDER_LAST_NAME", "")
EMAIL      = os.environ.get("SENDER_EMAIL", "")
PHONE      = os.environ.get("SENDER_PHONE", "")

MESSAGE = open("contact.txt", encoding="utf-8").read().strip()

XLSX_FILE  = "namur_private_sellers.xlsx"
DELAY_MIN  = 2.5
DELAY_MAX  = 5.0


# ── Helpers ───────────────────────────────────────────────────────────────────
def ts():
    return datetime.now().strftime("%H:%M:%S")


def sleep(lo=DELAY_MIN, hi=DELAY_MAX):
    time.sleep(random.uniform(lo, hi))


def sleep_progress(seconds: float, label: str = "cooldown"):
    bar_width = 30
    start = time.time()
    end   = start + seconds
    while True:
        elapsed   = time.time() - start
        ratio     = min(elapsed / seconds, 1.0)
        filled    = int(bar_width * ratio)
        bar       = "█" * filled + "░" * (bar_width - filled)
        remaining = max(0, seconds - elapsed)
        sys.stdout.write(f"\r  [{ts()}] {label}  [{bar}] {remaining:.0f}s remaining  ")
        sys.stdout.flush()
        if time.time() >= end:
            break
        time.sleep(0.5)
    sys.stdout.write("\n")


def fetch(page: Page, url: str) -> bool:
    """Navigate to url with 403/429 retry logic. Returns True on success."""
    for attempt in range(1, 4):
        try:
            response = page.goto(url, wait_until="domcontentloaded", timeout=40_000)
            status   = response.status if response else 0

            if status == 200:
                return True
            elif status == 429:
                wait = random.uniform(45, 90)
                sleep_progress(wait, f"WARN  Rate limited (429)")
            elif status == 403:
                wait = random.uniform(60, 120) * attempt
                sleep_progress(wait, f"WARN  403 attempt {attempt}/3")
            else:
                print(f"  [{ts()}] FAIL  HTTP {status} — {url}")
                return False

        except Exception as e:
            wait = 10 * attempt
            print(f"  [{ts()}] WARN  Error attempt {attempt}/3: {e} — sleeping {wait}s")
            if attempt == 3:
                return False
            time.sleep(wait)

    print(f"  [{ts()}] FAIL  gave up on {url}")
    return False


def accept_cookies(page: Page):
    try:
        btn = page.locator(
            "button#didomi-notice-agree-button, "
            "button:has-text('OK'), "
            "button:has-text('Accepter'), "
            "button:has-text('Accept all'), "
            "button:has-text('Tout accepter')"
        ).first
        btn.wait_for(timeout=8_000)
        btn.click()
        time.sleep(1.5)
    except PWTimeout:
        pass


def fill_contact_form(page: Page, url: str) -> bool:
    """Navigate to a listing, fill and submit the contact form. Returns True on success."""
    if not fetch(page, url):
        return False

    accept_cookies(page)

    # Zoom out so the full dialog fits on screen
    page.evaluate("document.body.style.zoom = '0.75'")

    # Scroll gently after cookies dismissed to appear human
    page.evaluate("window.scrollTo(0, document.body.scrollHeight * 0.4)")
    sleep(1.5, 2.5)

    # ── Click "Get in touch" to open the modal ────────────────────────────────
    page.evaluate("window.scrollTo(0, document.body.scrollHeight * 0.6)")
    sleep(0.8, 1.2)
    try:
        get_in_touch = page.locator(
            "button:has-text('Get in touch'), "
            "button:has-text('Prendre contact'), "
            "button:has-text('Contacter')"
        ).first
        get_in_touch.wait_for(timeout=10_000)
        get_in_touch.scroll_into_view_if_needed()
        get_in_touch.click()
        sleep(1.0, 2.0)
    except PWTimeout:
        print(f"  [{ts()}] FAIL  'Get in touch' button not found")
        return False

    # ── Wait for dialog then fill fields scoped inside it ────────────────────
    dialog = page.get_by_role("dialog", name="Get in touch")
    try:
        dialog.wait_for(state="visible", timeout=8_000)
    except PWTimeout:
        print(f"  [{ts()}] FAIL  dialog did not open")
        return False

    try:
        dialog.locator("input[name='firstName']").fill(FIRST_NAME)
        dialog.locator("input[name='lastName']").fill(LAST_NAME)
        dialog.locator("input[name='email']").fill(EMAIL)
        dialog.locator("input[name='phone']").fill(PHONE)
        # Select "No" for "I already own a property" — click the label
        no_id = dialog.locator("input[id*='sellingProperty-no']").get_attribute("id")
        dialog.locator(f"label[for='{no_id}']").click()
        # Check "Ask for more info"
        ask_id = dialog.locator("input[id*='askMoreInfo']").get_attribute("id")
        dialog.locator(f"label[for='{ask_id}']").click()
        dialog.locator("textarea[name='message']").fill(MESSAGE)
        sleep(1.0, 2.0)
    except PWTimeout as e:
        print(f"  [{ts()}] FAIL  form field not found: {e}")
        return False

    # ── Click submit and wait for dialog to close ────────────────────────────
    try:
        submit_btn = dialog.locator("button.sendMyRequestInModalButton")
        submit_btn.wait_for(state="attached", timeout=5_000)
        submit_btn.scroll_into_view_if_needed()
        sleep(0.5, 1.0)
        submit_btn.click()
    except PWTimeout:
        print(f"  [{ts()}] FAIL  submit button not found")
        return False

    # Success = dialog closes OR shows a success message inside
    try:
        page.locator(
            "text=Request successfully sent, "
            "text=successfully sent, "
            "text=Demande envoyée, "
            "text=envoyée avec succès, "
            "text=Votre message a été envoyé"
        ).first.wait_for(timeout=10_000)
        return True
    except PWTimeout:
        pass
    try:
        dialog.wait_for(state="hidden", timeout=5_000)
        return True
    except PWTimeout:
        print(f"  [{ts()}] FAIL  no success confirmation detected")
        return False


# ── Main ──────────────────────────────────────────────────────────────────────
def main(limit=None):
    wb = openpyxl.load_workbook(XLSX_FILE)
    ws = wb.active

    header = [cell.value for cell in ws[1]]
    url_col  = header.index("url")  + 1
    done_col = header.index("done") + 1

    rows = [
        (r, ws.cell(r, url_col).value)
        for r in range(2, ws.max_row + 1)
        if not ws.cell(r, done_col).value
    ]
    if limit:
        rows = rows[:limit]

    total   = len(rows)
    success = 0
    failed  = 0

    print(f"[{ts()}] {total} listings to contact\n")

    PROFILE_DIR = os.path.join(os.path.dirname(__file__), "chrome_profile")

    with sync_playwright() as p:
        context = p.chromium.launch_persistent_context(
            user_data_dir=PROFILE_DIR,
            channel="chrome",
            headless=False,
            args=[
                "--disable-blink-features=AutomationControlled",
                "--no-sandbox",
                "--disable-dev-shm-usage",
                "--no-first-run",
                "--no-default-browser-check",
                "--disable-infobars",
            ],
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            locale="fr-BE",
            viewport={"width": 1280, "height": 800},
        )
        context.add_init_script(
            "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
        )
        page = context.new_page()

        for idx, (row_num, url) in enumerate(rows, 1):
            print(f"  [{ts()}] [{idx}/{total}] {url}")
            ok = fill_contact_form(page, url)

            if ok:
                ws.cell(row_num, done_col).value = "done"
                wb.save(XLSX_FILE)
                success += 1
                print(f"  [{ts()}]   ✓ sent — progress saved")
            else:
                ws.cell(row_num, done_col).value = "failed"
                wb.save(XLSX_FILE)
                failed += 1
                print(f"  [{ts()}]   ✗ failed — marked in xlsx")

            # Cooldown every 15 listings
            if idx % 15 == 0:
                wait = random.uniform(60, 120)
                sleep_progress(wait, f"cooldown after {idx} listings")
            else:
                sleep()

        try:
            context.close()
        except Exception:
            pass

    print(f"\n[{ts()}] Done — {success} sent, {failed} failed out of {total}")


if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("--test", action="store_true", help="Process only the first listing")
    args = ap.parse_args()
    main(limit=1 if args.test else None)
