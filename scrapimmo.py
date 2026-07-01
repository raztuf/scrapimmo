import json
import os
import random
import time
import openpyxl
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright, Page

from common import (
    PROVINCES, DELAY_MIN, DELAY_MAX, BROWSER_ARGS, USER_AGENT, LOCALE, VIEWPORT,
    WEBDRIVER_INIT_SCRIPT, ts, sleep_progress, fetch, extract_json_object,
)

# ── Config ────────────────────────────────────────────────────────────────────
COLUMNS = ["url", "postal_code", "locality", "price", "done"]


def choose_province() -> dict:
    print("Select province to scrape:")
    for key, p in PROVINCES.items():
        print(f"  {key}. {p['name']}")
    while True:
        choice = input("Enter number: ").strip()
        if choice in PROVINCES:
            return PROVINCES[choice]
        print("Invalid choice, try again.")


def is_target(postal_code: str, ranges: list[tuple[int, int]]) -> bool:
    try:
        code = int(postal_code.strip())
        return any(lo <= code <= hi for lo, hi in ranges)
    except ValueError:
        return False


# ── Search URL builder ────────────────────────────────────────────────────────
BASE_URL     = "https://www.immoweb.be/en/search/{type}/for-sale/{slug}?countries=BE&page={page}&orderBy=newest"
SEARCH_TYPES = ("house", "apartment")


def parse_price(listing) -> str:
    """Read the clean price from the <iw-price> JSON attribute.

    The visible text is rendered by a Vue component that concatenates two spans
    (e.g. '€259,000259000€'), so we read the structured ':price' attribute
    instead — 'mainDisplayPrice' is the formatted string, 'mainValue' the number.
    """
    iw = listing.find("iw-price")
    if iw and iw.get(":price"):
        try:
            data = json.loads(iw.get(":price"))
        except json.JSONDecodeError:
            data = None
        if data:
            disp = data.get("mainDisplayPrice")
            if disp:
                return disp.strip()
            mv = data.get("mainValue")
            if mv is not None:
                return f"€{int(mv):,}"
    # Fallback: rendered text (may be mangled, but better than nothing)
    price_tag = listing.find("p", class_="card--result__price")
    return price_tag.get_text(strip=True) if price_tag else ""


# ── Parse a single search result page ─────────────────────────────────────────
def parse_page(html: str, postal_ranges: list[tuple[int, int]]) -> list[dict]:
    soup     = BeautifulSoup(html, "html.parser")
    listings = soup.find_all("article", class_="card--result")
    results  = []

    for listing in listings:
        # ── Filter: skip agency listings ──────────────────────────────────────
        if listing.find("img", src=lambda s: s and "/customers/" in s):
            continue

        # ── Extract URL ────────────────────────────────────────────────────────
        link = listing.find("a", class_="card__title-link")
        if not link or not link.get("href"):
            continue

        # ── Extract locality, filter by target range ───────────────────────────
        locality_tag  = listing.find("p", class_="card--results__information--locality")
        locality_text = locality_tag.get_text(strip=True) if locality_tag else ""
        parts         = locality_text.split(maxsplit=1)
        postal_code   = parts[0] if len(parts) >= 1 else ""
        locality      = parts[1] if len(parts) >= 2 else ""

        if not is_target(postal_code, postal_ranges):
            continue

        # ── Extract price ──────────────────────────────────────────────────────
        price = parse_price(listing)

        results.append({
            "url":         link["href"],
            "postal_code": postal_code,
            "locality":    locality,
            "price":       price,
        })

    return results


def is_private_owner(page: Page, listing_url: str) -> bool:
    print(f"    [{ts()}] CHECK {listing_url}")
    html = fetch(page, listing_url)
    time.sleep(random.uniform(DELAY_MIN, DELAY_MAX))
    if html is None:
        return False

    data = extract_json_object(html, "window.classified")
    if data is None:
        print(f"    [{ts()}]   SKIP could not parse window.classified")
        return False

    customers = data.get("customers") or []
    result    = bool(customers) and all(c.get("type") == "PRIVATE" for c in customers)
    print(f"    [{ts()}] {'  OK ✓ private owner' if result else '  SKIP agency/unknown'}")
    return result


# ── Excel helpers ──────────────────────────────────────────────────────────────
def load_existing_urls(filepath: str) -> set[str]:
    if not os.path.exists(filepath):
        return set()
    wb     = load_workbook(filepath)
    ws     = wb.active
    header = [cell.value for cell in ws[1]]
    if "url" not in header:
        return set()
    url_col = header.index("url") + 1
    return {ws.cell(row=r, column=url_col).value for r in range(2, ws.max_row + 1)}


def append_to_xlsx(results: list[dict], filepath: str):
    if not results:
        return
    if not os.path.exists(filepath):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(COLUMNS)
    else:
        wb = load_workbook(filepath)
        ws = wb.active
    for r in results:
        ws.append([r.get(c, "") for c in COLUMNS])
    wb.save(filepath)
    print(f"  [{ts()}] SAVED {len(results)} new row(s) → '{filepath}'")


# ── Main scraper ───────────────────────────────────────────────────────────────
def scrape():
    province    = choose_province()

    while True:
        raw = input("Start at page (default 1): ").strip()
        if raw == "":
            start_page = 1
            break
        if raw.isdigit() and int(raw) >= 1:
            start_page = int(raw)
            break
        print("Invalid, enter a number >= 1.")

    output_file = province["output_file"]
    slug        = province["url_slug"]
    postal_ranges = province["postal_ranges"]
    prov_name   = province["name"]

    seen_urls = load_existing_urls(output_file)
    if seen_urls:
        print(f"[{ts()}] Loaded {len(seen_urls)} existing URLs — will skip these\n")

    total_new = 0
    print(f"[{ts()}] Scraping {prov_name} — {len(SEARCH_TYPES)} types, page 1 onwards until empty\n")

    with sync_playwright() as p:
        print(f"[{ts()}] Launching Chrome...")
        browser = p.chromium.launch(channel="chrome", headless=False, args=BROWSER_ARGS)
        print(f"[{ts()}] Chrome launched, opening page...")
        context = browser.new_context(
            user_agent=USER_AGENT,
            locale=LOCALE,
            viewport=VIEWPORT,
        )
        context.add_init_script(WEBDRIVER_INIT_SCRIPT)
        page = context.new_page()
        print(f"[{ts()}] Page ready, starting scrape...")

        for search_type in SEARCH_TYPES:
            print(f"[{ts()}] ── {search_type.upper()} ──────────────────────────────────")
            pg = start_page - 1
            while True:
                pg += 1
                url = BASE_URL.format(type=search_type, slug=slug, page=pg)
                print(f"[{ts()}] Page {pg:>3}  fetching search results...")

                # ── Cooldown every 10 pages ───────────────────────────────────
                if pg % 10 == 0:
                    cooldown = random.uniform(30, 60)
                    sleep_progress(cooldown, f"[{ts()}] Page {pg:>3}  cooldown")

                try:
                    html = fetch(page, url)
                    if html is None:
                        time.sleep(random.uniform(DELAY_MIN, DELAY_MAX))
                        continue

                    # Random scroll to appear human
                    page.evaluate("window.scrollTo(0, Math.random() * document.body.scrollHeight)")
                    time.sleep(random.uniform(0.5, 1.5))

                    soup = BeautifulSoup(html, "html.parser")
                    if not soup.find_all("article", class_="card--result"):
                        print(f"[{ts()}] Page {pg:>3}  no listings found — done")
                        break

                    candidates     = parse_page(html, postal_ranges)
                    new_candidates = [r for r in candidates if r["url"] not in seen_urls]
                    seen_urls.update(r["url"] for r in new_candidates)

                    print(f"[{ts()}] Page {pg:>3}  {len(candidates)} {prov_name} candidates ({len(new_candidates)} new) — verifying each...")

                    confirmed = []
                    for r in new_candidates:
                        if is_private_owner(page, r["url"]):
                            confirmed.append(r)

                    append_to_xlsx(confirmed, output_file)
                    total_new += len(confirmed)
                    print(f"[{ts()}] Page {pg:>3}  confirmed {len(confirmed)}/{len(new_candidates)} private owners  (running total: {total_new})\n")

                except Exception as e:
                    print(f"[{ts()}] ERROR {url} → {e}")

                time.sleep(random.uniform(DELAY_MIN, DELAY_MAX))

        context.close()
        browser.close()

    print(f"[{ts()}] Done — {total_new} new private owner listings appended to '{output_file}'")


# ── Run ────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    scrape()
